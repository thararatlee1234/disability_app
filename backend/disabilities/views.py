from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Count, Q
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from .models import MedicalCheck, PersonWithDisability
from .serializers import (
    CaseInsensitiveTokenObtainPairSerializer,
    MedicalCheckSerializer,
    PersonSerializer,
)

def index(request):
    return redirect('/static/web/index.html')

class CaseInsensitiveTokenObtainPairView(TokenObtainPairView):
    serializer_class = CaseInsensitiveTokenObtainPairSerializer


class PersonViewSet(viewsets.ModelViewSet):
    queryset = PersonWithDisability.objects.all().order_by('first_name','last_name').prefetch_related('medical_checks')
    serializer_class = PersonSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset().filter(owner=self.request.user)
        
        citizen_id = self.request.query_params.get('citizen_id', '').strip()
        if citizen_id:
            qs = qs.filter(citizen_id=citizen_id)

        # Advanced Address Filters
        province = self.request.query_params.get('province')
        if province:
            qs = qs.filter(province__icontains=province)
            
        district = self.request.query_params.get('district')
        if district:
            qs = qs.filter(district__icontains=district)
            
        subdistrict = self.request.query_params.get('subdistrict')
        if subdistrict:
            qs = qs.filter(subdistrict__icontains=subdistrict)

        house_no = self.request.query_params.get('house_no')
        if house_no:
            qs = qs.filter(house_no__icontains=house_no)

        village_no = self.request.query_params.get('village_no')
        if village_no:
            qs = qs.filter(village_no=village_no)

        address = self.request.query_params.get('address')
        if address:
            qs = qs.filter(address__icontains=address)

        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search) | 
                Q(last_name__icontains=search) | 
                Q(citizen_id__icontains=search) | 
                Q(phone__icontains=search)
            )
        return qs

    def list(self, request, *args, **kwargs):
        if request.query_params.get('all') == 'true':
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        return super().list(request, *args, **kwargs)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)


class MedicalCheckViewSet(viewsets.ModelViewSet):
    queryset = MedicalCheck.objects.select_related('person').prefetch_related('photos')
    serializer_class = MedicalCheckSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset().filter(person__owner=self.request.user)
        person_id = self.request.query_params.get('person')
        if person_id:
            qs = qs.filter(person_id=person_id)
        print(f"DEBUG: Fetching Checks for user {self.request.user}. Count: {qs.count()}")
        return qs

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def check_report(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not start_date or not end_date:
        return Response({'detail': 'กรุณาระบุ start_date และ end_date (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Base queryset for persons owned by the user
    persons = PersonWithDisability.objects.filter(owner=request.user)
    
    # People who have a medical check in the range
    checked_persons = persons.filter(
        medical_checks__check_date__range=[start_date, end_date]
    ).distinct().order_by('first_name', 'last_name')
    
    # People who do NOT have a medical check in the range
    unchecked_persons = persons.exclude(
        id__in=checked_persons.values_list('id', flat=True)
    ).order_by('first_name', 'last_name')
    
    # Use context in serializer for absolute URLs of photos
    context = {'request': request}
    
    return Response({
        'checked': PersonSerializer(checked_persons, many=True, context=context).data,
        'unchecked': PersonSerializer(unchecked_persons, many=True, context=context).data,
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_check_report(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not start_date or not end_date:
        return Response({'detail': 'กรุณาระบุ start_date และ end_date (YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
    
    persons = PersonWithDisability.objects.filter(owner=request.user)
    
    checked_persons = persons.filter(
        medical_checks__check_date__range=[start_date, end_date]
    ).distinct().order_by('first_name', 'last_name')
    
    unchecked_persons = persons.exclude(
        id__in=checked_persons.values_list('id', flat=True)
    ).order_by('first_name', 'last_name')

    wb = Workbook()
    
    # Sheet 1: Checked
    ws1 = wb.active
    ws1.title = "ตรวจแล้ว"
    headers = ['ลำดับ', 'ชื่อ-นามสกุล', 'เลขบัตรประชาชน', 'เบอร์โทรศัพท์', 'ประเภทความพิการ', 'ตำบล', 'อำเภอ', 'จังหวัด', 'วันที่ตรวจในรอบนี้', 'รายละเอียดการตรวจ']
    ws1.append(headers)
    for i, p in enumerate(checked_persons, 1):
        # Get checks for this person in the range
        checks_in_range = p.medical_checks.filter(check_date__range=[start_date, end_date]).order_by('check_date')
        dates_str = ", ".join([c.check_date.strftime('%Y-%m-%d') for c in checks_in_range])
        details_str = " | ".join([c.detail for c in checks_in_range])
        
        ws1.append([i, p.full_name, p.citizen_id, p.phone, p.disability_type, p.subdistrict, p.district, p.province, dates_str, details_str])

    # Sheet 2: Unchecked
    ws2 = wb.create_sheet(title="ยังไม่ได้ตรวจ")
    ws2.append(headers)
    for i, p in enumerate(unchecked_persons, 1):
        ws2.append([i, p.full_name, p.citizen_id, p.phone, p.disability_type, p.subdistrict, p.district, p.province])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="check_report_{start_date}_to_{end_date}.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def stats(request):
    qs = PersonWithDisability.objects.filter(owner=request.user)
    return Response({
        'total': qs.count(),
        'by_province': list(qs.values('province').annotate(total=Count('id')).order_by('-total')[:30]),
        'by_disability_type': list(qs.values('disability_type').annotate(total=Count('id')).order_by('-total')[:50]),
    })

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel_view(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'detail': 'กรุณาแนบไฟล์ Excel'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from .importer import import_persons_from_excel
        result = import_persons_from_excel(file, request.user)
        return Response(result)
    except Exception as e:
        return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def download_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = [
        'เลขบัตรประชาชน', 'คำนำหน้า', 'ชื่อ', 'นามสกุล', 'เพศ',
        'ประเภทความพิการ', 'เบอร์โทรศัพท์', 'ที่อยู่ (บ้านเลขที่ หมู่ ถนน)', 
        'ตำบล', 'อำเภอ', 'จังหวัด', 'ละติจูด', 'ลองจิจูด', 'หมายเหตุ'
    ]
    ws.append(headers)
    
    # Add an example row
    ws.append([
        '1234567890123', 'นาย', 'สมชาย', 'ใจดี', 'ชาย',
        'พิการทางการเคลื่อนไหว', '0812345678', '123 ม.1', 'ในเมือง', 'เมือง',
        'กรุงเทพฯ', '13.7563', '100.5018', 'ตัวอย่างข้อมูล'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="template_disability.xlsx"'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_excel_view(request):
    qs = PersonWithDisability.objects.filter(owner=request.user).order_by('first_name','last_name')
    wb = Workbook()
    ws = wb.active
    ws.title = "Disability Report"
    headers = [
        'เลขบัตรประชาชน', 'คำนำหน้า', 'ชื่อ', 'นามสกุล', 'เพศ',
        'ประเภทความพิการ', 'เบอร์โทรศัพท์', 'ที่อยู่', 'ตำบล', 'อำเภอ', 'จังหวัด',
        'ละติจูด', 'ลองจิจูด', 'หมายเหตุ'
    ]
    ws.append(headers)
    for p in qs:
        ws.append([
            p.citizen_id, p.prefix, p.first_name, p.last_name, p.gender,
            p.disability_type, p.phone, p.address, p.subdistrict, p.district, p.province,
            p.latitude, p.longitude, p.notes
        ])
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="disability_report_full.xlsx"'
    wb.save(response)
    return response
