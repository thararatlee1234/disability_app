import re
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from .models import PersonWithDisability

PREFIXES = ['นาย','นาง','นางสาว','ด.ช.','ด.ญ.','เด็กชาย','เด็กหญิง']

def clean(v):
    return '' if v is None else str(v).strip()

def normalize_key(s):
    return re.sub(r'\s+', '', clean(s).lower())

def pick(row, headers, names, exclude=None):
    exclude = [normalize_key(e) for e in (exclude or [])]
    normalized_names = [normalize_key(n) for n in names]
    
    # Try exact match first
    for idx, h in headers.items():
        if any(h == n for n in normalized_names):
            if exclude and any(e in h for e in exclude):
                continue
            return clean(row[idx])
            
    # Try partial match if exact failed
    is_looking_for_fullname = any('สกุล' in n for n in normalized_names)
    
    for n in normalized_names:
        for idx, h in headers.items():
            if n in h:
                if not is_looking_for_fullname and 'สกุล' in h and n == 'ชื่อ':
                    continue
                if exclude and any(e in h for e in exclude):
                    continue
                return clean(row[idx])
    return ''

def parse_name(name):
    name = clean(name)
    prefix = ''
    for p in sorted(PREFIXES, key=len, reverse=True):
        if name.startswith(p):
            prefix = p
            name = name[len(p):].strip()
            break
    parts = name.split()
    if not parts:
        return prefix, '', ''
    f = parts[0]
    l = ' '.join(parts[1:]) if len(parts) > 1 else ''
    return prefix, f, l

def dec(v):
    try:
        if v is None: return None
        s = str(v).strip().replace(',', '')
        return Decimal(s) if s else None
    except (InvalidOperation, ValueError):
        return None

def parse_address_fields(addr_str):
    """
    Parses a Thai address string into components safely.
    """
    res = {
        'house_no': '',
        'village_no': '',
        'village_name': '',
        'road': '',
        'subdistrict': '',
        'district': '',
        'province': '',
        'postal_code': ''
    }
    if not addr_str:
        return res
    
    # Pre-clean
    addr_str = re.sub(r'^ที่อยู่\s*', '', addr_str).strip()
    
    # Extract Postal Code
    pc_match = re.search(r'(\d{5})$', addr_str)
    if pc_match:
        res['postal_code'] = pc_match.group(1)
        addr_str = addr_str[:pc_match.start()].strip()
        
    # Extract Province
    prov_match = re.search(r'(?:จ\.|จังหวัด)\s*([^\s]+)', addr_str)
    if prov_match:
        res['province'] = prov_match.group(1).strip()
        addr_str = addr_str[:prov_match.start()].strip()
        
    # Extract District
    dist_match = re.search(r'(?:อ\.|อำเภอ)\s*([^\s]+)', addr_str)
    if dist_match:
        res['district'] = dist_match.group(1).strip()
        addr_str = addr_str[:dist_match.start()].strip()
        
    # Extract Subdistrict
    sub_match = re.search(r'(?:ต\.|ตำบล)\s*([^\s]+)', addr_str)
    if sub_match:
        res['subdistrict'] = sub_match.group(1).strip()
        addr_str = addr_str[:sub_match.start()].strip()
        
    # Extract Road
    road_match = re.search(r'(?:ถ\.|ถนน)\s*([^\s]+)', addr_str)
    if road_match:
        res['road'] = road_match.group(1).strip()
        addr_str = addr_str[:road_match.start()].strip()
        
    # Extract Village No
    vn_match = re.search(r'(?:ม\.|หมู่ที่)\s*([^\s]+)', addr_str)
    if vn_match:
        res['village_no'] = vn_match.group(1).strip()
        addr_str = addr_str[:vn_match.start()].strip()
        
    # Village Name
    vname_match = re.search(r'หมู่บ้าน\s*([^\s]+)', addr_str)
    if vname_match:
        res['village_name'] = vname_match.group(1).strip()
        addr_str = addr_str[:vname_match.start()].strip()
        
    # Remainder is House No
    res['house_no'] = addr_str.strip()
    
    return res

def import_excel(file_or_path, sheet_name=None, exclude_columns=None, owner=None):
    if exclude_columns is None:
        exclude_columns = []
    exclude_columns_norm = [normalize_key(c) for c in exclude_columns]

    wb = load_workbook(file_or_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active
        
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created':0, 'updated':0, 'skipped':0}
        
    # Detect header row
    header_row_idx = 0
    for i, row in enumerate(rows[:20]): # Check first 20 rows
        row_vals = [normalize_key(str(v)) for v in row if v]
        if any(k in row_vals for k in ['ชื่อ','นามสกุล','เลขบัตร','เลขประจำตัวประชาชน','citizenid','ชื่อ-นามสกุล','ชื่อสกุล']):
            header_row_idx = i
            break
            
    header_values = rows[header_row_idx]
    headers = {i: normalize_key(str(v)) for i, v in enumerate(header_values) if v is not None}
    
    created = updated = skipped = 0
    for row_num, row in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):
        citizen_id = pick(row, headers, ['เลขบัตรประชาชน','บัตรประชาชน','เลขประจำตัวประชาชน','cid','เลขบัตร'])
        first_name = pick(row, headers, ['ชื่อ','firstname','first_name'])
        last_name = pick(row, headers, ['นามสกุล','lastname','last_name'])
        prefix = pick(row, headers, ['คำนำหน้า','คำนาหน้า','prefix'])
        full_name = pick(row, headers, ['ชื่อ-นามสกุล','ชื่อ นามสกุล','fullname','full_name','ชื่อสกุล'])
        
        if first_name == full_name:
            first_name = ''
            
        if not (first_name and last_name) and full_name:
            p, f, l = parse_name(full_name)
            prefix = prefix or p
            first_name = first_name or f
            last_name = last_name or l
            
        if not (first_name or last_name or citizen_id):
            skipped += 1
            continue
            
        raw_data = {}
        for i, v in enumerate(row):
            if i < len(header_values):
                orig_key = clean(header_values[i])
                if orig_key:
                    if normalize_key(orig_key) in exclude_columns_norm:
                        continue
                    raw_data[orig_key] = clean(v)

        address_raw = pick(row, headers, ['ที่อยู่','บ้านเลขที่','address'])
        subdistrict = pick(row, headers, ['ตำบล','ตําบล','subdistrict'])
        district = pick(row, headers, ['อำเภอ','อําเภอ','district'])
        province = pick(row, headers, ['จังหวัด','province'])
        map_url = pick(row, headers, ['ลิงก์','แผนที่','map','map_url'])
        
        addr_fields = parse_address_fields(address_raw)

        latitude = dec(pick(row, headers, ['latitude','lat','ละติจูด','gps','พิกัด']))
        longitude = dec(pick(row, headers, ['longitude','lng','lon','ลองจิจูด','gps','พิกัด']))

        # If both are in the same 'gps' or 'พิกัด' column, try to split them
        if latitude is not None and longitude is None:
             val = str(latitude)
             if ',' in val:
                 parts = val.split(',')
                 latitude = dec(parts[0])
                 longitude = dec(parts[1])

        # Try to extract from map_url if still missing
        if (latitude is None or longitude is None) and map_url:
            # 1. Prioritize !3d (lat) and !4d (lng)
            m_pin = re.search(r'!3d([-\d.]+)!4d([-\d.]+)', map_url)
            if m_pin:
                latitude = latitude or dec(m_pin.group(1))
                longitude = longitude or dec(m_pin.group(2))
            
            if latitude is None:
                # 2. Look for @lat,lng
                m_at = re.search(r'@([-\d.]+),([-\d.]+)', map_url)
                if m_at:
                    latitude = latitude or dec(m_at.group(1))
                    longitude = longitude or dec(m_at.group(2))

            if latitude is None:
                # 3. Look for query=lat,lng or q=lat,lng or ll=lat,lng
                m_q = re.search(r'[?&](?:query|q|ll)=([-\d.]+),([-\d.]+)', map_url)
                if m_q:
                    latitude = latitude or dec(m_q.group(1))
                    longitude = longitude or dec(m_q.group(2))

            if latitude is None:
                # 4. Look for place/lat,lng or search/lat,lng
                m_place = re.search(r'(?:place|search)/([-\d.]+),([-\d.]+)', map_url)
                if m_place:
                    latitude = latitude or dec(m_place.group(1))
                    longitude = longitude or dec(m_place.group(2))

        payload = dict(
            prefix=prefix,
            first_name=first_name,
            last_name=last_name,
            gender=pick(row, headers, ['เพศ','gender']),
            disability_type=pick(row, headers, ['ประเภทความพิการ','ความพิการ','disability','อาการ']),
            phone=pick(row, headers, ['โทรศัพท์','เบอร์','phone','tel','เบอร์โทร']),
            address=address_raw,
            house_no=addr_fields['house_no'],
            village_no=addr_fields['village_no'],
            village_name=addr_fields['village_name'],
            road=addr_fields['road'],
            subdistrict=subdistrict or addr_fields['subdistrict'],
            district=district or addr_fields['district'],
            province=province or addr_fields['province'],
            postal_code=addr_fields['postal_code'],
            map_url=map_url,
            notes=clean(pick(row, headers, ['หมายเหตุ','notes'])),
            latitude=latitude,
            longitude=longitude,
            source_row=row_num,
            raw_data=raw_data,
        )
        
        if owner is not None:
            payload['owner'] = owner

        if citizen_id:
            lookup = {'citizen_id': citizen_id}
            if owner is not None:
                lookup['owner'] = owner
            obj, was_created = PersonWithDisability.objects.update_or_create(**lookup, defaults=payload)
        else:
            lookup = {'first_name': first_name, 'last_name': last_name, 'source_row': row_num}
            if owner is not None:
                lookup['owner'] = owner
            obj, was_created = PersonWithDisability.objects.update_or_create(**lookup, defaults={**payload, 'citizen_id': citizen_id})
        
        created += 1 if was_created else 0
        updated += 0 if was_created else 1
        
    return {'created':created, 'updated':updated, 'skipped':skipped}
